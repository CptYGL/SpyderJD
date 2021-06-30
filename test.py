# -*- coding:utf-8 -*-
from __future__ import print_function
import re
from openpyxl import load_workbook, Workbook
from pyecharts import options as opts
from pyecharts.charts import Bar,Line
from pyecharts.globals import ThemeType
#from matplotlib import pyplot as plt
import jieba
from pyecharts.options.global_options import AxisOpts
'''
调用pyecharts在网页显示灵动的图表
将prefilt处理成postfilt,一百条评论合一做词云,总评论(销量)和价格做散点图(好评越多点越大),
#这里是写入postfilt.slsx
    try:
        data.remove('postfilt')
        data.save(filename)
    except:
        print('还没有post表,即将创建...\n')
    post_sheet = data.worksheets[1]
    post_sheet.title = 'postfilt'
    for subli in sum:
        post_sheet.append(subli)
        data.save(filename)
'''
#filter 过滤器含多个子列表,把手机型号对应的所有项目不重样的放在一个子列表,写给postfilt.xslx(没啥用,给老师看的)
def filter(filename):
    data = load_workbook(filename)
    pref_sheet = data.worksheets[0]
    list_all = []
    sum = []
    #excel转换成列表
    for m in pref_sheet.rows:
        list_row = []
        for n in m:
            list_row.append(n.value)#gbk2utf(str(n.value)))
        list_all.append(list_row)
    #设置一个计数器prod_cnt计算到底多少个产品(因为爬虫不一定准确爬到100条,有的产品不一定够十页)
    prod_cnt = list(set([x[0] for x in list_all]))                          #这个列表里是不重样的ID(当然也可以改成别的)
    for i in range(len(prod_cnt)):
        tmp_str = '空,'
        tmp_li = []
        for line in range(len(list_all)):
            if list_all[line][0]==prod_cnt[i]:                              #在listall遍历,找出相同的,然后该列除了
                tmp_str += list_all[line][7]                                #[7]即'评论'外都给tmplist,tmpstr来加总相同
                tmp_li += list_all[line][:7]                                #产品的评论,该ID检索完后,拼成新列表给sum
        tmp_str=tmp_str.replace('\n',',').replace(':',',').replace(' ',',')
        tmp_str=re.sub('[a-zA-Z0-9]','',tmp_str)                            #去掉一些没用的字符
        a = list(set(tmp_li))
        a.sort(key=tmp_li.index)
        a.append(tmp_str)
        sum.append(a)
    for i in range(len(sum)):
        if len(sum[i])<8:sum[i]=sum[i+1]
    return sum
#analyzer 做统计词频
def analyzer(filted):
    counts = []
    #keywds是因为感觉jieba经常出一些奇怪的词,为了筛选添加的关键词
    keywds = ['快速','清晰','绚丽','一流','舒适','上乘','黑色','超棒','美丽',\
        '流畅','舒适','牌面','美观','漂亮','颜色','耐用','面子','好看','完美',\
        '好玩','运动','不错','很好','前卫','酷炫','发热','着魔','神器','顺手',\
        '大气','实惠','智能','强大','郁闷','遗憾','勉强','难看','便宜','太贵']
    #我懒,不弄了,需要的话再用
    for item in filted:
        count = {}
        words = jieba.lcut(str(item[7]))
        for word in words:
            if (len(word)>=3)or(word in keywds):count[word] = count.get(word,0)+1
            else:continue
        counts.append(list(count.items()))
    for i in range(len(counts)):
        counts[i].sort(key=lambda x:x[1],reverse=True)
    return [x[0] for x in counts]
#getter 返回一个字典 {key='特征',value=[]}
def getter(filted):
    name = [x[1] for x in filted]
    for i in range(len(name)):
        name[i]='-'.join(name[i].split()[:4])
    price = [x[2] for x in filted]
    vol = [x[3] for x in filted]
    c_good = [x[4] for x in filted]
    c_mid = [x[5] for x in filted]
    c_bad = [x[6] for x in filted]
    impres1 = analyzer(filted)
    dic = {'name':name,'price':price,'volume':vol,'goodcom':c_good,'mid-com':c_mid,\
        'bad-com':c_bad,'impression1':impres1}
    return dic
#画图函数pyecharts
def overlap_bar_line(name,price,vol,good,mid,bad,imp):
    bar = (
        Bar(init_opts=opts.InitOpts(theme=ThemeType.DARK))
        .add_xaxis(name)
        .add_yaxis('销量',vol,stack='stack1',category_gap='50%')
        .add_yaxis('好评',good,stack='stack2',category_gap='50%')
        .add_yaxis('中评',mid,stack='stack2',category_gap='50%')
        .add_yaxis('差评',bad,stack='stack2',category_gap='50%')
        .extend_axis(
            yaxis=opts.AxisOpts(
                axislabel_opts=opts.LabelOpts(formatter="{value}元"), interval=1000
            )
        )
        #.set_series_opts(markpoint_opts=opts.MarkPointOpts(data=imp))
        .set_global_opts(
            title_opts=opts.TitleOpts(title='销量-价格-评价'),
            yaxis_opts=opts.AxisOpts(
                axislabel_opts=opts.LabelOpts(formatter='{value}台')
            ),
            datazoom_opts=opts.DataZoomOpts(
                    orient='horizontal'

            )
        )
        #.set_series_opts(label_opts=opts.LabelOpts(position="right"))
    )
    line = Bar()
    line.add_xaxis(name).add_yaxis('价格',price,yaxis_index=1)
    bar.overlap(line).render('test.html')


a = filter('data.xlsx')
dict_sum = getter(a)
overlap_bar_line(dict_sum['name'],dict_sum['price'],dict_sum['volume'],dict_sum['goodcom'],dict_sum['mid-com'],dict_sum['bad-com'],dict_sum['impression1'])
#写入excel
'''
data = load_workbook('post.xlsx')
sheet = data.worksheets[0]
for i in range(len(a)):
    l =[str(dict_sum['name'][i]),str(dict_sum['price'][i]),str(dict_sum['volume'][i]),str(dict_sum['goodcom'][i]),str(dict_sum['mid-com'][i]),str(dict_sum['bad-com'][i]),str(dict_sum['impression1'][i])]
    sheet.append(l)
data.save('post.xlsx')
'''